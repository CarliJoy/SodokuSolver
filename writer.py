from pathlib import Path
from typing import List
from openpyxl import load_workbook, worksheet
from openpyxl.styles import Font


def write_result(input_filename: Path, result: List[List[int]], output_filename: Path):
    """
    Load an example, fill it with the solution and save it as a different file
    """
    # Load Workbook
    wb = load_workbook(input_filename)
    # Get first sheet
    sh: worksheet = wb[wb.sheetnames[0]]
    for i, line in enumerate(result, start=1):
        for j, val in enumerate(line, start=1):
            cell_value = sh.cell(i, j).value
            if cell_value is None:
                sh.cell(i, j).value = int(val)
                sh.cell(i, j).font = sh.cell(i, j).font + Font(
                    color="9A9A9A", italic=True
                )
            elif int(cell_value) != int(val):
                print(f"!! Warning  Input cells do not match!"
                      f" ({i}, {j}) {cell_value} != {val}")
    wb.save(output_filename)
